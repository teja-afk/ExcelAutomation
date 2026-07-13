
from openpyxl import load_workbook

try:

    SOURCE_FILE = "source.xlsx"
    TEMPLATE_FILE = "template.xlsx"

    SOURCE_SHEET = 0      # First sheet
    TEMPLATE_SHEET = 0    # First sheet

    START_ROW = 2         # Row 1 = headers

    # Load workbooks
    src_wb = load_workbook(SOURCE_FILE, data_only=True)
    tmp_wb = load_workbook(TEMPLATE_FILE)

    src_ws = src_wb.worksheets[SOURCE_SHEET]
    tmp_ws = tmp_wb.worksheets[TEMPLATE_SHEET]

    # Optional: Clear previous values in B:H
    for row in range(START_ROW, tmp_ws.max_row + 1):
        for col in range(2, 9):      # B:H
            tmp_ws.cell(row=row, column=col).value = None

    # Copy values
    for row in range(START_ROW, src_ws.max_row + 1):
        for col in range(2, 9):      # B:H
            tmp_ws.cell(row=row, column=col).value = src_ws.cell(row=row, column=col).value

    # Save
    tmp_wb.save(TEMPLATE_FILE)

    print("Data copied successfully.")
except FileNotFoundError as e:
    print(f"File not found: {e}")
except Exception as e:
    print(f"Error: {e}")
